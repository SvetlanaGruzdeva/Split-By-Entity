# SplitByEntity - splits one consolidated file in separate files, one file per entity

import openpyxl, os, logging
from pprint import pprint
from tkinter.filedialog import askopenfilename

logging.basicConfig(filename='logs.txt', level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
logging.disable(logging.CRITICAL)

columnList = []
entitiesList = []
consolFile = askopenfilename() # Selected by user from browser
region = input('Type in region: ')

if region not in consolFile:
    print('Wrong region name, programm has been stopped.')
    exit()

wb = openpyxl.load_workbook(consolFile, data_only=True, keep_vba=True) # data_only - to show cells value, not formula
ws = wb.get_sheet_by_name('Summary')
column = ws['D']
columnList = [column[x].value for x in range(5, len(column))]
# logging.debug(columnList)

for value in columnList:
    if value not in entitiesList:
        entitiesList.append(value)
entitiesList.remove('FORMULA')
logging.debug('List of unique entities as per consolidated file: ')
logging.debug(entitiesList)

logging.info('Started entity loop.')
for entity in entitiesList:
    logging.debug('Entity code once loop just started: ' + entity)
    
    logging.debug('Opening consolidated file ' + consolFile)
    wb = openpyxl.load_workbook(consolFile, data_only=True, keep_vba=True)
    ws = wb.get_sheet_by_name('Summary')
    
    logging.info('Deleting lines...')
    for rowNum in range (ws.max_row, 5, -1):
        logging.debug('Entity code when DELETE loop started: ' + entity)
        if ws.cell(row=rowNum, column=4).value != entity and ws.cell(row=rowNum, column=4).value != 'FORMULA':
            logging.debug('Row number is ' + str(rowNum))
            logging.debug('Row value is ' + str(ws.cell(row=rowNum, column=4).value) + ' and entity value is ' + entity)
            ws.delete_rows(rowNum)
    
    logging.info('Deleting is finished.')
    logging.info('Entity loop finished.')    

    ws.protection.password = '1234'
    ws.protection.sheet = True
    ws.protection.autoFilter = False
    
    if region == 'KZ':
        insctuctions = 'Инструкции'
    else:
        insctuctions = 'Instructions'
    # TODO: Unselect summary tab first
    wb.active = wb.sheetnames.index(insctuctions)
    wb['Summary'].views.sheetView[0].tabSelected = False

    newFileName = consolFile.replace(region, entity)
    wb.save(newFileName)

